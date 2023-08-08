import tkinter as tk
import os
import json
import xlwings as xw
from tkinter import filedialog, ttk, simpledialog
from datetime import datetime, timedelta
from typing import List, Dict

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle

from src.models import TaskType, StaffMember, Task, WorkDay, Shift
from src.staff import SKILLSET_MAP

input_file_paths = []
input_schedule_days = []
input_roster = ""


def create_window():
    root = tk.Tk()
    root.title("Study Schedule Selection")
    root.geometry("500x400")  # Set the initial size of the window (width x height)

    file_select_button = tk.Button(root, text="Select Schedule", command=open_file_dialog)
    file_select_button.pack(pady=10)

    roster_select_button = tk.Button(root, text="Select Roster", command=open_roster_dialog)
    roster_select_button.pack(pady=10)

    global file_table
    columns = ("Study Schedule", "Day")
    file_table = ttk.Treeview(root, columns=columns, show="headings", selectmode="browse")

    # Set column headings
    for col in columns:
        file_table.heading(col, text=col)
        file_table.column(col, width=150, anchor="center")

    file_table.pack(pady=5)

    generate_schedule_button = tk.Button(root, text="Generate Schedules", command=output_directory_dialog)
    generate_schedule_button.pack(pady=10)

    root.mainloop()


def open_file_dialog():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    input_file_paths.append(file_path)

    file_name = file_path.split("/")[-1]  # Get only the filename from the path

    # Prompt the user to enter the day number using a dialog box
    day = simpledialog.askinteger("Enter Day", f"Enter the day for '{file_name}':")
    input_schedule_days.append(day)

    # Add the file name and the entered number to the table
    file_table.insert("", "end", values=(file_name, day))


def open_roster_dialog():
    input_roster = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    day = simpledialog.askstring("Enter Day", f"Enter the schedule date: e.g. MON 07 AUG")

    print(input_roster)

    get_staff_shifts(input_roster, day)


def output_directory_dialog():
    output_file_path = filedialog.askdirectory()
    copy_schedule_days(input_file_paths, output_file_path)


def find_token_row_index(token, df):
    return df[df.apply(lambda row: row.astype(str).str.contains(token, case=False).any(), axis=1)].index


def append_empty_rows(sheet, num_rows):
    for _ in range(num_rows):
        empty_row = [None] * sheet.max_column
        sheet.append(empty_row)


def copy_schedule_days(input_files, output_file_directory):
    # Create a new workbook for the output if it doesn't exist
    output_file_name = os.path.join(output_file_directory, "Daily_Schedules.xlsx")
    if not os.path.exists(output_file_name):
        output_wb = Workbook()
        output_wb.save(output_file_name)
    else:
        output_wb = openpyxl.load_workbook(output_file_name, values_only=True)
    
    i = 0

    for input_file in input_files:

        file_name = input_file.split("/")[-1]

        study_name = file_name.split()[0]

        df = pd.read_excel(input_file)

        start_token = "Day " + str(input_schedule_days[i])
        end_token = "Workbook"

        start_index = find_token_row_index(start_token, df)
        end_index = find_token_row_index(end_token, df)

        if not start_index.empty and not end_index.empty:
            start_index = start_index[0]
            
            # Find the first occurrence of the ending token after the start token
            end_index = end_index[end_index > start_index][0]

            # Create a DataFrame for the study name row
            name_row = [[study_name] * len(df.columns)]
            name_row_df = pd.DataFrame(name_row, columns=df.columns)

            # Extract the section of data between the start token and the end token
            selected_data = df.loc[start_index : end_index, df.columns != df.columns[0]]  # Exclude the first column from being copied

            # Combine the name row DataFrame with the selected_data DataFrame
            named_data = pd.concat([name_row_df, selected_data], ignore_index=True)

            # Get the selected sheet to copy data
            sheet = output_wb.active

            # Append the new data to the end of the sheet
            for row in dataframe_to_rows(selected_data, index=False, header=False):
                sheet.append(row)

            # Define a custom style for datetime values
            time_style = NamedStyle(name="time_style", number_format="HH:MM")

            # Formatting date time
            for row in sheet.iter_rows(min_row=sheet.max_row - len(selected_data) + 1, max_row=sheet.max_row):
                for cell in row:
                    source_cell = df.iloc[cell.row - sheet.max_row + len(selected_data) - 1, cell.column - 1]

                     # Check if the value is a datetime object and apply the custom style
                    if isinstance(source_cell, pd.Timestamp):
                        cell.style = time_style
                    
                    #preserve the row height
                    sheet.row_dimensions[cell.row].height = sheet.row_dimensions[cell.row - sheet.max_row + len(selected_data)].height

            # Preserve column width
            for column_cells in sheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                sheet.column_dimensions[column_cells[0].column_letter].width = length + 2  # Additional padding

            output_wb.save(output_file_name)

        else:
            print("Day not found in study schedule")
        
        append_empty_rows(sheet, 4)

        i=i+1

        print(f"Successfully created Daily_schedules.xslx")

        excel_to_json(output_file_name, output_file_directory)

        # TODO(liv): 
        #   Generate a list of Staff members for the day (of type StaffMember)
        #   Generate a list of tasks for the relevant studies (of type Task)
        #   Put that all into a WorkDay object
        # # Might look something like:

        # staff_members: List[StaffMember] = get_staff_members(...)
        # tasks: List[Task] = get_tasks(...)
        # work_day = WorkDay(
        #     tasks=tasks,
        #     staff_members=staff_members,
        # )

        # # save to json
        # with open("outputs/work_day.json", "w") as f:
        #     f.write(work_day.json(indent=4))
        

def get_staff_shifts(roster_path: str, day: str):

    shift_ranges = {
        'morning_short': ['D15:D25'],               #for shifts 0730-1330
        'morning_long': ['H15:H25', 'L15:L25'],     #for shifts 0700-1500    
        'morning_late': ['P15:P25'],                #for shifts 0800-1600
        'afternoon': ['H27:H32', 'L27:L32'],        #for shifts 1430-2230
        'night': ['H34:H37', 'L34:L37']             #for shifts 2200-0730
    }

    # Load the workbook
    app = xw.App(visible=False)
    wb = app.books.open(roster_path, password='nzcr')
    sheet = wb.sheets[day]
    
    rostered = {}
    
    for category, column_ranges in shift_ranges.items():
        data = []
        for column_range in column_ranges:
            column_cells = sheet[column_range]
            for cell in column_cells:
                cell_value = cell.value
                if cell_value is not None:  # Skip empty cells
                    data.append(cell_value.split()[0])
        rostered[category] = data

    wb.close()
    app.quit()

    for category, data in rostered.items():
        print(f"{category}:", data)

    current_date = datetime.now()
    date = datetime.strptime(day, "%a %d %b").replace(year=current_date.year)
    
    
    get_staff_members(rostered, date)


def get_staff_members(rostered: Dict[str, List[str]], roster_date: datetime):

    shift_types = {
        'morning_short' : Shift(start_time=roster_date.replace(hour=7, minute=30), finish_time=roster_date.replace(hour=13, minute=30)),
        'morning_long' : Shift(start_time=roster_date.replace(hour=7, minute=00), finish_time=roster_date.replace(hour=15, minute=00)),
        'morning_late' : Shift(start_time=roster_date.replace(hour=8, minute=00), finish_time=roster_date.replace(hour=16, minute=0)),
        'afternoon' : Shift(start_time=roster_date.replace(hour=14, minute=30), finish_time=roster_date.replace(hour=22, minute=30)),
        'night' : Shift(start_time=roster_date.replace(hour=22, minute=00), finish_time=(roster_date + timedelta(days=1)).replace(hour=7, minute=30)),
    }

    staff_members = []

    for shifts in rostered:
        for current_staff in rostered[shifts]:
            staff_members.append(StaffMember(name=current_staff, shift=shift_types[shifts], attributes=SKILLSET_MAP[current_staff]))

    print(staff_members)


def assign_category(task_name):
    # Enumerate task names based on keywords

    assign_tasks = {
        TaskType.DOCTOR: ["EXAM", "EXAMINATION", "PHYSICAL", "DOCTOR", "ELIGIBILITY"],
        TaskType.NURSE: [
            " PLACEBO",
            "DOSING",
            "AE CHECK",
            "CHECK AE",
            "ADVERSE",
            "MEDICATION",
            "SITE",
            "ACCOUNTABILITY",
            "DISPENSE",
            "REPORTED",
            "PHONE"
        ],
        TaskType.CRT: [
            "SUBJECT",
            "STANDARDISED",
            "STANDARDIZED",
            "BREAKFAST",
            "LUNCH",
            "DINNER",
            "SNACK",
            "MEAL",
            "SUPINE",
            "SEMI",
            "ECG",
            "VITAL",
            "VITALS",
            "STATS",
            "SATS ",
            "COMMENCE",
            "BLOCK",
            "PREG",
            "DOA",
            "URINALYSIS",
            "URINE",
            "MSU",
            "WEIGHT",
            "HEIGHT",
            "BMI",
            "ALCOHOL",
            "CLEAN",
            "LINEN",
            "WALK"
        ],
        TaskType.ANY: [
            "ADMIT",
            "WRIST",
            "WRISTBAND",
            "COMPLIANCE",
            "REMIND",
            "ENSURE",
            "RULES",
            "TOUR",
            "WORKBOOK",
            "DISCHARGE",
            "ENCOURAGE",
            "ORIENTATION",
            "RESTRICTIONS",
        ],
        TaskType.PHLEBOTOMY: ["BLOOD", "SAFETIES", "FLUSH"],
        TaskType.CANNULATION: ["CANNULATION"],
        TaskType.INFUSION: ["INF ", "IV ", "INTRAVENOUS"],
        TaskType.SPIRO: ["SPIROMETRY"],
        TaskType.SPUTUM: ["SPUTUM"],
        TaskType.BREEZING: ["REE ASSESSMENT"],
        TaskType.TRIPLICATE: ["TRIPLICATE"],
        TaskType.PHARMACY: ["RANDOMISATION", "RANDOMIZATION"],
        TaskType.DATE: ["DATE"],
        TaskType.DAY: ["DAY"]
    }

    for assignment in assign_tasks:
        for keyword in assign_tasks[assignment]:
            if keyword in str.upper(task_name):
                return assignment
        
    return TaskType.OTHER

    #TODO: hierachy of availability


def time_to_str(t):
    # Convert time object to a string in HH:MM format
    return t.strftime('%H:%M') if isinstance(t, time) else str(t)


def excel_to_json(schedules, json_file_directory):
    # Read the Excel file into a pandas DataFrame
    df = pd.read_excel(schedules, header=None)
    
    # Create a dictionary to store tasks by category
    data = {}

    current_category = None
    current_study = "Unknown"
    
    for row in df.itertuples(index=False, name=None):
        task = str(row[0])  # Access the first element of the tuple as the task name
        
        # Extract the study name from the task name if "Schedule" is present
        current_study = str(row[0]).split()[0]
        
        # Convert times to strings and skip NaN values
        times = [time_to_str(t) for t in row[1:] if pd.notna(t)]
        
        # Skip adding to data if task is NaN
        if task == "nan":
           continue

        current_category = assign_category(task).value
        
        # Create an instance of the appropriate Enum class based on the category
        if current_study not in data:
            data[current_study] = {}
        
        if current_category not in data[current_study]:
            data[current_study][current_category] = []
        
        # Append the task and times to the list for the category
        data[current_study][current_category].append({"task": task, "times": times})

    # Convert Enum keys to strings before writing to JSON
    json_data = {str(key): value for key, value in data.items()}

    json_file_path = os.path.join(json_file_directory, "JSON_schedules.json")

    # Create the JSON file and write the dictionary into it
    with open(json_file_path, 'w') as f:
        json.dump(json_data, f, indent=4)

    print(f"Created JSON file")


if __name__ == "__main__":
    create_window()
