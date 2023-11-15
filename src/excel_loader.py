import tkinter as tk
import os
import xlwings as xw
import datetime as dt
from tkinter import filedialog, ttk, simpledialog
from datetime import time, timedelta, datetime
from typing import List, Dict

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border

from models import TaskType, StaffMember, Task, WorkDay, Shift
from staff import SKILLSET_MAP


input_file_paths = []
input_cohort_names = []
input_schedule_days = []
staff_members_roster = []
study_patient_numbers = []      #TODO: patient dropout functionality?
date = "MON 07 AUG"
input_roster = ""
output_file_path = ""

#All subjects need column names (i.e. 1, 2, STBY etc.)
#All visits need to have Day x in the header
#All studies must be assigned a floor in the roster, same for staff
#If any rows in a schedule are not to be included, they should be deleted rather than minimized
#Tasks need to be clearly defined, i.e. "blood" needs to be included in "safeties"
#Triplicates need to be specified consecutively, i.e. not only "ECG triplicte 1 min apart" in one row

#TODO: determine task length, time spent in OP & what to do with returning staff

def create_window():
    root = tk.Tk()
    root.title("Study Schedule Selection")
    root.geometry("500x400")  # the initial size of the window (width x height)

    file_select_button = tk.Button(root, text="Select Schedule", command=open_file_dialog)
    file_select_button.pack(pady=10)

    roster_select_button = tk.Button(root, text="Select Roster", command=open_roster_dialog)
    roster_select_button.pack(pady=10)

    global file_table
    columns = ("Study Schedule", "Day")
    file_table = ttk.Treeview(root, columns=columns, show="headings", selectmode="browse")

    # Set column headings
    for col in columns:
        file_table.heading(col, text=col)
        file_table.column(col, width=150, anchor="center")

    file_table.pack(pady=5)

    generate_schedule_button = tk.Button(root, text="Generate Schedules", command=output_directory_dialog)
    generate_schedule_button.pack(pady=10)

    root.mainloop()


def open_file_dialog():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    input_file_paths.append(file_path)

    file_name = file_path.split("/")[-1].split()[0]  # Get only the filename from the path

    #identify the study name by the cohort
    ask_cohort = simpledialog.askstring("Enter Cohort", f"Enter the cohort for '{file_name}':")
    cohort_name = file_name + " " + ask_cohort
    input_cohort_names.append(cohort_name)

    no_patients = simpledialog.askinteger("Enter Number of Subjects", f"Enter the number of subjects for '{cohort_name}':")
    study_patient_numbers.append(no_patients)

    # Prompt the user to enter the day number using a dialog box
    day = simpledialog.askinteger("Enter Day", f"Enter the day for '{cohort_name}':")
    input_schedule_days.append(day)

    # Add the file name and the entered number to the table
    file_table.insert("", "end", values=(cohort_name, day))


def open_roster_dialog():
    input_roster = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])

    date = simpledialog.askstring("Enter Day", f"Enter the schedule date: e.g. MON 07 AUG")

    print(input_roster)

    get_staff_shifts(input_roster, date)


def output_directory_dialog():
    output_file_path = filedialog.askdirectory()
    create_task_list(input_file_paths)


def find_token_row_index(token, df):
    return df[df.apply(lambda row: row.astype(str).str.contains(token, case=False).any(), axis=1)].index


def append_empty_rows(sheet, num_rows):
    for _ in range(num_rows):
        empty_row = [None] * sheet.max_column
        sheet.append(empty_row)


def isTriplicate(task_name):
    if "triplicate" in str(task_name).strip().lower():
        return True
    else:
        return False


def create_task_list(input_file_paths):

    i = 0
    end_token = "Workbook"
    tasks = []
    current_date = datetime.now()
    roster_day = datetime.strptime(date, "%a %d %b").replace(year=current_date.year)

    for file in input_file_paths:
        
        date_count = 1
        schedule = {}
        patients = []

        current_study = input_cohort_names[i]
        current_floor = ""
        for key, study_array in study_floors.items():
            if any(study.strip().lower() in current_study.strip().lower() for study in study_array):
                current_floor = key
                break

        current_day = input_schedule_days[i]

        pt_token = "Day " + str(current_day)

        duration = 0
        fit = 0

        df = pd.read_excel(file)

        start_index = find_token_row_index(pt_token, df)
        end_index = find_token_row_index(end_token, df)

        if not start_index.empty and not end_index.empty:
            start_index = start_index[0]
            # Find the first occurrence of the ending token after the start token
            end_index = end_index[end_index > start_index][0]

            # Extract the section of data between the start token and the end token
            selected_data = df.loc[start_index : end_index, df.columns != df.columns[0]].copy()  # Exclude the first column from being copied
            print(selected_data)

        for row in selected_data.itertuples(index=False, name=None):
            first_cell = row[0] #taking strings from the first column
            if 'nan' in str(first_cell).strip().lower():
                continue
            if 'date' in str(first_cell).strip().lower():
                date_count = date_count + 1
                continue
            
            if pt_token in str(first_cell) and date_count != 0: #take patients as the 2nd cell onwards, if Day i is read
                p = 0
                for cell_value in row[1:]:
                    if str(cell_value).strip().lower() == 'nurse' or p == study_patient_numbers[i]:
                        break
                    patients.append(cell_value)
                    p = p+1
                date_count = 0
            else:
                task_name = first_cell
                task_times = {}
                #iterate through each patient column, add as key for subsequent times
                for patient, time in zip(patients, row[1:]):
                    if patient is not None:
                        if str(time).strip().lower() == 'nan':
                            # task_times[patient] = ""
                            continue
                        if isinstance(time, datetime):
                            time = time.time()
                        task_times[patient] = datetime.combine(roster_day, time)
                    else:
                        continue
                
                #add a task and its corresponding patient-time dict to schedule dict
                schedule[task_name] = task_times
            
        for task_iter, patient_iter in schedule.items():
            attribute_list = assign_attribute(task_iter)
            attribute = attribute_list[0]
            task_duration = attribute_list[1]
            task_fit = attribute_list[2]
            is_triplicate = isTriplicate(task_iter)
            for patient_i, time_i in patient_iter.items():
                tasks.append(Task(study=current_study, floor=current_floor, patient=patient_i, time=time_i, duration=task_duration, strict_staffing=task_fit, required_attributes=attribute, title=task_iter, triplicate=is_triplicate))

        i = i+1
    
    create_workday(tasks)


def create_workday(task_list):

    workday = WorkDay(tasks=task_list, staff_members=staff_members_roster)

    json_file_path = os.path.join(output_file_path, "workday.json")

    with open(json_file_path, 'w') as json_file:
        json_file.write(workday.json(indent=4))
    
    copy_schedules()


def copy_schedules():
    # Create a new destination Excel workbook
    output_file_name = os.path.join(output_file_path, "Daily_Schedules.xlsx")
    if not os.path.exists(output_file_name):
        output_wb = Workbook()
        output_wb.save(output_file_name)
    else:
        output_wb = openpyxl.load_workbook(output_file_name, data_only=True)
    
    destination_sheet_name = 'Workday ' + date
    destination_sheet = output_wb.active
    destination_sheet.title = destination_sheet_name

    end_token = 'Workbook'   

    i=0

    custom_style = NamedStyle(name='custom_style')
    custom_style.number_format = 'hh:mm'
    output_wb.add_named_style(custom_style)

    current_row = 1

    # Iterate through multiple source Excel files
    for source_file_path in input_file_paths:
        source_file = openpyxl.load_workbook(source_file_path, data_only=True)
        source_sheet = source_file.active

        study_name_array = source_file_path.split("/")[-1].split()
        study_name = study_name_array[0] + " " + study_name_array[1] + " " + study_name_array[2] + " " + study_name_array[3]

        number_patients = study_patient_numbers[i]

        append_study_name = [study_name] + [None] * (destination_sheet.max_column - 1)
        empty_row = ["   "] + [None] * (destination_sheet.max_column - 1)

        for name_index, value in enumerate(append_study_name, start=1):
            current_cell = destination_sheet.cell(row=current_row, column=name_index, value=value)
            current_cell.font = Font(bold=True)
        
        current_row +=1
        
        for empty_index, value in enumerate(empty_row, start=1):
            destination_sheet.cell(row=current_row, column=empty_index, value=value)
            destination_sheet.row_dimensions[current_row].height = 15

        current_row +=1

        start_row = None
        end_row = None

        current_day = input_schedule_days[i]
        start_token = "Day " + str(current_day)

        found_start = False

        # iterating through rows to find start and end rows
        for row in source_sheet.iter_rows(min_row=1, max_row=source_sheet.max_row, min_col=2, max_col=2):  # Adjust column index to 2 (second column)
            cell_value = row[0].value
            if start_token in str(cell_value):
                if not found_start:
                    start_row = row[0].row
                    found_start = True
            elif found_start and end_token in str(cell_value):
                end_row = row[0].row
                break

        # Check if start and end rows were found in the source sheet
        if start_row is not None and end_row is not None:
            # Copy the section from the source to the destination
            for row_index in range(start_row, end_row + 1):
                for col_index, source_cell in enumerate(source_sheet.iter_cols(min_row=row_index, max_row=row_index, min_col=2, max_col=number_patients + 2)):  # Skip the first column
                    destination_cell = destination_sheet.cell(row=current_row, column=col_index + 1)
                    if type(source_cell[0].value) == dt.time:
                        destination_cell.value = source_cell[0].value
                        destination_cell.style = custom_style
                    elif type(source_cell[0].value) == dt.datetime:
                        destination_cell.value = source_cell[0].value.time()
                        destination_cell.style = custom_style
                    else:
                        destination_cell.value = source_cell[0].value
                    destination_cell.font = openpyxl.styles.Font(size=source_cell[0].font.size)
                    destination_cell.fill = openpyxl.styles.PatternFill(start_color=source_cell[0].fill.start_color, end_color=source_cell[0].fill.end_color, fill_type=source_cell[0].fill.fill_type)
                    destination_cell.alignment = openpyxl.styles.Alignment(horizontal=source_cell[0].alignment.horizontal, vertical=source_cell[0].alignment.vertical)

                    # Check if the source cell has bold font
                    if source_cell[0].font and source_cell[0].font.bold:
                        destination_cell.font = Font(bold=True)  # Apply bold font style

                    # Copy cell borders
                    if source_cell[0].border:
                        destination_cell.border = Border(
                            left=source_cell[0].border.left,
                            right=source_cell[0].border.right,
                            top=source_cell[0].border.top,
                            bottom=source_cell[0].border.bottom
                        )

                current_row += 1

            for col_index in range(1, destination_sheet.max_column + 1):
                if col_index == 1:
                    destination_sheet.column_dimensions[openpyxl.utils.get_column_letter(col_index)].width = source_sheet.column_dimensions["B"].width
                else:
                    destination_sheet.column_dimensions[openpyxl.utils.get_column_letter(col_index)].width = 8

            # Set row heights in the destination sheet based on whether they contain text or are empty
            for row_index in range(1, destination_sheet.max_row + 1):
                row_values = [cell.value for cell in destination_sheet[row_index]]
                if any(row_values):
                    destination_sheet.row_dimensions[row_index].height = 15
                else:
                    destination_sheet.row_dimensions[row_index].height = 2

        # Close the source file
        source_file.close()

        for _ in range(4):
            for index, value in enumerate(empty_row, start=1):
                destination_sheet.cell(row=current_row, column=index, value=value)
                destination_sheet.row_dimensions[current_row].height = 15
            current_row +=1

        i=i+1
    # Save the destination workbook
    output_wb.save("Daily_Schedules.xlsx")

    # Close the destination workbook
    output_wb.close()


        # TODO(liv): 
        #   Generate a list of Staff members for the day (of type StaffMember) - done
        #   Generate a list of tasks for the relevant studies (of type Task) - DONE
        #   Put that all into a WorkDay object
        # # Might look something like:

        # staff_members: List[StaffMember] = get_staff_members(...)
        # tasks: List[Task] = get_tasks(...)
        # work_day = WorkDay(
        #     tasks=tasks,
        #     staff_members=staff_members,
        # )

        # # save to json
        # with open("outputs/work_day.json", "w") as f:
        #     f.write(work_day.json(indent=4))
        

def get_staff_shifts(roster_path: str, day: str):       #might need to create second instance of OP staff after OP schedule complete?

    shifts = {
        'morning_short': ['D15:D25'],               #for shifts 0730-1330
        'morning_long': ['H15:H25', 'L15:L25'],     #for shifts 0700-1500
        'morning_late': ['P15:P25'],                #for shifts 0800-1600
        'afternoon': ['H27:H32', 'L27:L32'],        #for shifts 1430-2230
        'night': ['H34:H37', 'L34:L37']             #for shifts 2200-0730
    }

    staff_floor_ranges = {
        'outpatients': ['D15:D25'],
        'level_one': ['H15:H25', 'H27:H32', 'H34:H37'],
        'ground_floor': ['L15:L25', 'L27:L32', 'L34:L37'],
        'level_two': ['P15:P25']
    }

    study_floor_ranges = {
        'outpatients': ['D4:D11'],
        'level_one': ['H4:H11'],
        'level_two': ['L4:L11'],
        'ground_floor': ['P4:P11']
    }

    # Load the workbook
    app = xw.App(visible=False)
    wb = app.books.open(roster_path, password='nzcr')
    sheet = wb.sheets[day]
    
    rostered = {}
    staff_floors = {}
    global study_floors
    study_floors = {}
    
    for shift, shift_ranges in shifts.items():
        staff_shifts = []
        for shift_range in shift_ranges:
            shift_column_cells = sheet[shift_range]
            for shift_cell in shift_column_cells:
                shift_cell_value = shift_cell.value
                if shift_cell_value is not None:
                    staff_shifts.append(shift_cell_value.split()[0])
        rostered[shift] = staff_shifts

    for floor, floor_ranges in staff_floor_ranges.items():
        floors = []
        for floor_range in floor_ranges:
            staff_floor_column_cells = sheet[floor_range]
            for staff_floor_cell in staff_floor_column_cells:
                staff_floor_cell_value = staff_floor_cell.value
                if staff_floor_cell_value is not None:
                    floors.append(staff_floor_cell_value.split()[0])
        staff_floors[floor] = floors

    for category, column_ranges in study_floor_ranges.items():
        data = []
        for column_range in column_ranges:
            column_cells = sheet[column_range]
            for cell in column_cells:
                cell_value = cell.value
                if cell_value is not None:
                    data.append(cell_value.split()[0])
        study_floors[category] = data


    wb.close()
    app.quit()

    current_date = datetime.now()
    roster_date = datetime.strptime(day, "%a %d %b").replace(year=current_date.year)
    
    get_staff_members(rostered, staff_floors, roster_date)


def get_staff_members(rostered: Dict[str, List[str]], staff_floors: Dict[str, List[str]], roster_date: datetime):

    shift_types = {
        'morning_short' : Shift(start_time=roster_date.replace(hour=7, minute=30), finish_time=roster_date.replace(hour=13, minute=30)),
        'morning_long' : Shift(start_time=roster_date.replace(hour=7, minute=00), finish_time=roster_date.replace(hour=15, minute=00)),
        'morning_late' : Shift(start_time=roster_date.replace(hour=8, minute=00), finish_time=roster_date.replace(hour=16, minute=0)),
        'afternoon' : Shift(start_time=roster_date.replace(hour=14, minute=30), finish_time=roster_date.replace(hour=22, minute=30)),
        'night' : Shift(start_time=roster_date.replace(hour=22, minute=00), finish_time=(roster_date + timedelta(days=1)).replace(hour=7, minute=30)),
    }

    for shifts in rostered:
        for current_staff in rostered[shifts]:
            current_floor = next((key for key, staff_array in staff_floors.items() if current_staff in staff_array), None)
            staff_members_roster.append(StaffMember(name=current_staff, shift=shift_types[shifts], attributes=SKILLSET_MAP[current_staff], floor=current_floor))
    
    print(f"Roster complete")

def assign_attribute(task_name):
    # Enumerate task names based on keywords

    assign_tasks = {
        TaskType.DOCTOR: ["EXAM", "EXAMINATION", "PHYSICAL", "DOCTOR", "ELIGIBILITY"],
        TaskType.NURSE: [
            " PLACEBO",
            "DOSING",
            "AE CHECK",
            "CHECK AE",
            "ADVERSE",
            "MEDICATION",
            "SITE",
            "ACCOUNTABILITY",
            "DISPENSE",
            "REPORTED",
            "PHONE"
        ],
        TaskType.CRT: [
            "SUBJECT",
            "STANDARDISED",
            "STANDARDIZED",
            "BREAKFAST",
            "LUNCH",
            "DINNER",
            "SNACK",
            "MEAL",
            "SUPINE",
            "SEMI",
            "ECG",
            "VITAL",
            "VITALS",
            "STATS",
            "SATS ",
            "COMMENCE",
            "BLOCK",
            "PREG",
            "DOA",
            "URINALYSIS",
            "URINE",
            "MSU",
            "WEIGHT",
            "HEIGHT",
            "BMI",
            "ALCOHOL",
            "CLEAN",
            "LINEN",
            "WALK"
        ],
        TaskType.ANY: [
            "ADMIT",
            "WRIST",
            "WRISTBAND",
            "COMPLIANCE",
            "REMIND",
            "ENSURE",
            "RULES",
            "TOUR",
            "WORKBOOK",
            "DISCHARGE",
            "ENCOURAGE",
            "ORIENTATION",
            "RESTRICTIONS",
        ],
        TaskType.PHLEBOTOMY: ["BLOOD", "SAFETIES", "FLUSH"],
        TaskType.CANNULATION: ["CANNULATION"],
        TaskType.INFUSION: ["INF ", "IV ", "INTRAVENOUS"],
        TaskType.SPIRO: ["SPIROMETRY"],
        TaskType.SPUTUM: ["SPUTUM"],
        TaskType.BREEZING: ["REE ASSESSMENT"],
        TaskType.TRIPLICATE: ["TRIPLICATE"],
        TaskType.PHARMACY: ["RANDOMISATION", "RANDOMIZATION"],
        TaskType.DATE: ["DATE"],
        TaskType.DAY: ["DAY"]
    }

    current_task = TaskType.OTHER
    duration = 2
    fit = 0

    criteria = []

    for assignment in assign_tasks:
        for keyword in assign_tasks[assignment]:
            if keyword in str.upper(task_name):
                current_task = assignment

    if current_task == TaskType.DOCTOR:
        duration = 10
        fit = 1
    elif current_task == TaskType.CRT:
        duration = 3
        fit = 0
    elif current_task == TaskType.ANY:
        duration = 1
        fit = 0
    elif current_task == TaskType.PHLEBOTOMY:
        duration = 5
        fit = 1
    elif current_task == TaskType.CANNULATION:
        duration = 5
        fit = 1
    elif current_task == TaskType.INFUSION:
        duration = 5
        fit = 1
    elif current_task == TaskType.SPIRO:
        duration = 10
        fit = 1
    elif current_task == TaskType.SPUTUM:
        duration = 10
        fit = 1
    elif current_task == TaskType.BREEZING:
        duration = 5
        fit = 1
    elif current_task == TaskType.TRIPLICATE:
        duration = 1
        fit = 0
    elif current_task == TaskType.PHARMACY:
        duration = 10
        fit = 1
    
    criteria = [current_task, duration, fit]
        
    return criteria


def time_to_str(t):
    # Convert time object to a string in HH:MM format
    return t.strftime('%H:%M') if isinstance(t, time) else str(t)


if __name__ == "__main__":
    create_window()
