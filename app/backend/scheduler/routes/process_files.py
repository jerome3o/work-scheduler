from typing import Annotated, Dict, List
import openpyxl
import os
import pandas as pd
from scheduler.constants import PASSWORD
import msoffcrypto
from fastapi import APIRouter, File, Form, UploadFile
from datetime import time, timedelta, datetime
from scheduler.staff import SKILLSET_MAP

from io import BytesIO

from scheduler.models import (
    Task,
    TaskType,
    WorkDay,
    RosterProcessingResult,
    StudyScheduleProcessingResult,
    GenerateWorkDayOptions,
    StudyScheduleOption,
    Shift,
    StaffMember
)


router = APIRouter(
    prefix="/api/process-files",
    tags=["process-files"],
)


@router.post("/roster", response_model=RosterProcessingResult)
def process_roster(
    roster_file: UploadFile = Annotated[bytes, File()],
) -> RosterProcessingResult:
    # TODO(olivia): implement

    encrypted = BytesIO(roster_file.file.read())
    decrypted = decrypt_file(encrypted)

    wb = openpyxl.load_workbook(decrypted)
    print(wb.sheetnames)

    wb.close()

    return RosterProcessingResult(
        days=wb.sheetnames,
    )


@router.post("/study-schedule", response_model=StudyScheduleProcessingResult)
def process_study_schedule(
    study_schedule_file: UploadFile = File(...),
) -> StudyScheduleProcessingResult:

    f = BytesIO(study_schedule_file.file.read())

    wb = openpyxl.load_workbook(f)
    print(wb.sheetnames)

    return StudyScheduleProcessingResult(
        days=[],
        cohorts=[],
        patients=[],
    )


@router.post("/build-workday", response_model=WorkDay)
def build_workday(
    study_schedule_files: list[UploadFile] = File(...),
    roster_file: UploadFile = File(...),
    additional_data: str = Form(...),
) -> WorkDay:
    options = GenerateWorkDayOptions.model_validate_json(additional_data)

    #debugging the schedule file list
    for filename in study_schedule_files:
        print(f"{filename.filename}")

    #------------Roster Processing------------#

    encrypted = BytesIO(roster_file.file.read())
    decrypted = decrypt_file(encrypted)

    wb = openpyxl.load_workbook(decrypted)

    day = options.roster_day

    roster = get_staff_shifts(wb, day)
    staff = roster[0]
    staff_floor = roster[1]
    study_floors = roster[2]

    current_date = datetime.now()
    roster_date = datetime.strptime(day, "%a %d %b").replace(year=current_date.year)
    
    staff_list = get_staff_members(staff, staff_floor, roster_date)
    
    #------------Study Schedule Processing------------#

    task_list = []

    for schedule_file, study_schedule_option in zip(study_schedule_files, options.study_schedule_options):
        schedule_name = schedule_file.filename    
        task_list.extend(create_task_list(schedule_file, study_schedule_option, study_floors, schedule_name, roster_date))

    #------------Workday Generation------------#

    #print(study_schedule_files)
    #print(roster_file)
    #print(options)
    return WorkDay(
        tasks=task_list,
        staff_members=staff_list,
        relative_to=roster_date,
    )


def create_workday(task_list: List[Task], staff_list: List[StaffMember], roster_date: datetime, output_file_path: str):

    workday = WorkDay(tasks=task_list, staff_members=staff_list, relative_to=roster_date)

    json_file_path = os.path.join(output_file_path, "workday.json")

    with open(json_file_path, 'w') as json_file:
        json_file.write(workday.model_dump_json(indent=4))


def decrypt_file(encrypted: BytesIO) -> BytesIO:
    decrypted = BytesIO()
    file = msoffcrypto.OfficeFile(encrypted)

    # TODO: the password needs to be input
    file.load_key(password=PASSWORD)
    file.decrypt(decrypted)

    return decrypted


def isTriplicate(task_name):
    if "triplicate" in str(task_name).strip().lower():
        return True
    else:
        return False


def get_staff_shifts(roster: openpyxl.Workbook, day: str):       #might need to create second instance of OP staff after OP schedule complete?

    shifts = {
        'morning_short': ['D15:D25'],               #for shifts 0730-1330
        'morning_long': ['H15:H25', 'L15:L25'],     #for shifts 0700-1500
        'morning_late': ['P15:P25'],                #for shifts 0800-1600
        'afternoon': ['H27:H32', 'L27:L32'],        #for shifts 1430-2230
        'night': ['H34:H37', 'L34:L37']             #for shifts 2200-0730
    }

    staff_floor_ranges = {
        'outpatients': ['D15:D25'],
        'level_one': ['H15:H25', 'H27:H32', 'H34:H37'],
        'ground_floor': ['L15:L25', 'L27:L32', 'L34:L37'],
        'level_two': ['P15:P25']
    }

    study_floor_ranges = {
        'outpatients': ['D4:D11'],
        'level_one': ['H4:H11'],
        'level_two': ['L4:L11'],
        'ground_floor': ['P4:P11']
    }

    # Load the workbook
    sheet = roster[day]

    rostered = {}
    staff_floors = {}
    study_floors = {}
    
    for shift, shift_ranges in shifts.items():
        staff_shifts = []
        for shift_range in shift_ranges:
            shift_column_cells = sheet[shift_range]
            for shift_cell in shift_column_cells:
                shift_cell_value = shift_cell[0].value
                if shift_cell_value is not None:
                    staff_shifts.append(shift_cell_value.split()[0])
        rostered[shift] = staff_shifts

    for floor, floor_ranges in staff_floor_ranges.items():
        floors = []
        for floor_range in floor_ranges:
            staff_floor_column_cells = sheet[floor_range]
            for staff_floor_cell in staff_floor_column_cells:
                staff_floor_cell_value = staff_floor_cell[0].value
                if staff_floor_cell_value is not None:
                    floors.append(staff_floor_cell_value.split()[0])
        staff_floors[floor] = floors
    
    for category, column_ranges in study_floor_ranges.items():
        data = []
        for column_range in column_ranges:
            column_cells = sheet[column_range]
            for cell in column_cells:
                cell_value = cell[0].value
                if cell_value is not None:
                    data.append(cell_value.split()[0])
        study_floors[category] = data

    return rostered, staff_floors, study_floors


def get_staff_members(rostered: Dict[str, List[str]], staff_floors: Dict[str, List[str]], roster_date: datetime):

    staff_members_roster = []

    shift_types = {
        'morning_short' : Shift(start_time=roster_date.replace(hour=7, minute=30), finish_time=roster_date.replace(hour=13, minute=30)),
        'morning_long' : Shift(start_time=roster_date.replace(hour=7, minute=00), finish_time=roster_date.replace(hour=15, minute=00)),
        'morning_late' : Shift(start_time=roster_date.replace(hour=8, minute=00), finish_time=roster_date.replace(hour=16, minute=0)),
        'afternoon' : Shift(start_time=roster_date.replace(hour=14, minute=30), finish_time=roster_date.replace(hour=22, minute=30)),
        'night' : Shift(start_time=roster_date.replace(hour=22, minute=00), finish_time=(roster_date + timedelta(days=1)).replace(hour=7, minute=30)),
    }

    for shifts in rostered:
        for current_staff in rostered[shifts]:
            current_floor = next((key for key, staff_array in staff_floors.items() if current_staff in staff_array), None)
            staff_members_roster.append(StaffMember(name=current_staff, shift=shift_types[shifts], attributes=SKILLSET_MAP[current_staff], floor=current_floor))
    
    return staff_members_roster


def find_token_row_index(token: str, df: pd.DataFrame) -> pd.Index:
    return df[df.apply(lambda row: row.astype(str).str.contains(token, case=False).any(), axis=1)].index


def create_task_list(study_schedule_files: BytesIO , options: StudyScheduleOption, study_floors: Dict[str, List[str]], schedule_name: str, roster_date: datetime):

    tasks = []

    schedule = BytesIO(study_schedule_files.file.read())
    wb = openpyxl.load_workbook(schedule)
    
    date_count = 1
    processed_schedule = {}
    patients = []
    current_floor = ""

    for key, study_array in study_floors.items():
        if any(study.strip().lower() in schedule_name.strip().lower() for study in study_array):
            current_floor = key
            break
    
    current_day = options.day
    pt_token = "Day " + str(current_day)
    end_token = "Workbook"
    duration = 0
    fit = 0

    df = pd.read_excel(schedule)

    start_index = find_token_row_index(pt_token, df)
    end_index = find_token_row_index(end_token, df)

    if not start_index.empty and not end_index.empty:
        start_index = start_index[0]
        # Find the first occurrence of the ending token after the start token
        end_index = end_index[end_index > start_index][0]
        # Extract the section of data between the start token and the end token
        selected_data = df.loc[start_index : end_index, df.columns != df.columns[0]].copy()  # Exclude the first column from being copied
        print(selected_data)

    for row in selected_data.itertuples(index=False, name=None):

        first_cell = row[0] #taking strings from the first column

        if 'nan' in str(first_cell).strip().lower():
            continue
        if 'date' in str(first_cell).strip().lower():
            date_count = date_count + 1
            continue
        
        if pt_token in str(first_cell) and date_count != 0: #take patients as the 2nd cell onwards, if Day i is read
            p = 0
            for cell_value in row[1:]:
                if str(cell_value).strip().lower() == 'nurse' or p == options.patients:
                    break
                patients.append(cell_value)
                p = p+1
            date_count = 0
        else:
            task_name = first_cell
            task_times = {}

            #iterate through each patient column, add as key for subsequent times
            for patient, time in zip(patients, row[1:]):
                if patient is not None:
                    if str(time).strip().lower() == 'nan':
                        # task_times[patient] = ""
                        continue
                    if isinstance(time, datetime):
                        time = time.time()
                    task_times[patient] = datetime.combine(roster_date, time)
                else:
                    continue
            
            #add a task and its corresponding patient-time dict to schedule dict
            processed_schedule[task_name] = task_times
        
    for task_iter, patient_iter in processed_schedule.items():
        attribute_list = assign_attribute(task_iter)
        attribute = attribute_list[0]
        task_duration = attribute_list[1]
        task_fit = attribute_list[2]
        is_triplicate = isTriplicate(task_iter)

        for patient_i, time_i in patient_iter.items():
            tasks.append(Task(study=schedule_name, floor=current_floor, patient=patient_i, time=time_i, duration=task_duration, strict_staffing=task_fit, required_attributes=attribute, title=task_iter, triplicate=is_triplicate))

    return tasks


def assign_attribute(task_name):
    # Enumerate task names based on keywords

    assign_tasks = {
        TaskType.DOCTOR: ["EXAM", "EXAMINATION", "PHYSICAL", "DOCTOR", "ELIGIBILITY"],
        TaskType.NURSE: [
            " PLACEBO",
            "DOSING",
            "AE CHECK",
            "CHECK AE",
            "ADVERSE",
            "MEDICATION",
            "SITE",
            "ACCOUNTABILITY",
            "DISPENSE",
            "REPORTED",
            "PHONE"
        ],
        TaskType.CRT: [
            "SUBJECT",
            "STANDARDISED",
            "STANDARDIZED",
            "BREAKFAST",
            "LUNCH",
            "DINNER",
            "SNACK",
            "MEAL",
            "SUPINE",
            "SEMI",
            "ECG",
            "VITAL",
            "VITALS",
            "STATS",
            "SATS ",
            "COMMENCE",
            "BLOCK",
            "PREG",
            "DOA",
            "URINALYSIS",
            "URINE",
            "MSU",
            "WEIGHT",
            "HEIGHT",
            "BMI",
            "ALCOHOL",
            "CLEAN",
            "LINEN",
            "WALK"
        ],
        TaskType.ANY: [
            "ADMIT",
            "WRIST",
            "WRISTBAND",
            "COMPLIANCE",
            "REMIND",
            "ENSURE",
            "RULES",
            "TOUR",
            "WORKBOOK",
            "DISCHARGE",
            "ENCOURAGE",
            "ORIENTATION",
            "RESTRICTIONS",
        ],
        TaskType.PHLEBOTOMY: ["BLOOD", "SAFETIES", "FLUSH"],
        TaskType.CANNULATION: ["CANNULATION"],
        TaskType.INFUSION: ["INF ", "IV ", "INTRAVENOUS"],
        TaskType.SPIRO: ["SPIROMETRY"],
        TaskType.SPUTUM: ["SPUTUM"],
        TaskType.BREEZING: ["REE ASSESSMENT"],
        TaskType.TRIPLICATE: ["TRIPLICATE"],
        TaskType.PHARMACY: ["RANDOMISATION", "RANDOMIZATION"],
        TaskType.DATE: ["DATE"],
        TaskType.DAY: ["DAY"]
    }

    current_task = TaskType.OTHER
    duration = 2
    fit = 0

    criteria = []

    for assignment in assign_tasks:
        for keyword in assign_tasks[assignment]:
            if keyword in str.upper(task_name):
                current_task = assignment

    if current_task == TaskType.DOCTOR:
        duration = 10
        fit = 1
    elif current_task == TaskType.CRT:
        duration = 3
        fit = 0
    elif current_task == TaskType.ANY:
        duration = 1
        fit = 0
    elif current_task == TaskType.PHLEBOTOMY:
        duration = 5
        fit = 1
    elif current_task == TaskType.CANNULATION:
        duration = 5
        fit = 1
    elif current_task == TaskType.INFUSION:
        duration = 5
        fit = 1
    elif current_task == TaskType.SPIRO:
        duration = 10
        fit = 1
    elif current_task == TaskType.SPUTUM:
        duration = 10
        fit = 1
    elif current_task == TaskType.BREEZING:
        duration = 5
        fit = 1
    elif current_task == TaskType.TRIPLICATE:
        duration = 1
        fit = 0
    elif current_task == TaskType.PHARMACY:
        duration = 10
        fit = 1
    
    criteria = [current_task, duration, fit]
        
    return criteria